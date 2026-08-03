"""Oylik va davomat hisobotlarini Excel (.xlsx) ga chiqarish."""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import attendance as att_mod
from app.models import Attendance, DayStatus, Employee
from app.payroll import PayrollResult
from app.timeutil import UZ_WEEKDAYS, fmt_time, iter_days, month_bounds

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TOTAL_FONT = Font(bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY_FMT = "#,##0"

_STATUS_LABELS = {
    DayStatus.PRESENT: "Vaqtida",
    DayStatus.LATE: "Kechikdi",
    DayStatus.ABSENT: "Kelmadi",
    DayStatus.LEAVE: "Ta'til/ruxsat",
    DayStatus.HOLIDAY: "Dam olish",
    DayStatus.INCOMPLETE: "Ketishi yo'q",
}

_STATUS_FILLS = {
    DayStatus.LATE: PatternFill("solid", fgColor="FFF2CC"),
    DayStatus.ABSENT: PatternFill("solid", fgColor="FFC7CE"),
    DayStatus.INCOMPLETE: PatternFill("solid", fgColor="FCE4D6"),
    DayStatus.HOLIDAY: PatternFill("solid", fgColor="EDEDED"),
    DayStatus.LEAVE: PatternFill("solid", fgColor="DDEBF7"),
}


def _write_header(sheet, row: int, titles: list[str]) -> None:
    for col, title in enumerate(titles, start=1):
        cell = sheet.cell(row=row, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    sheet.row_dimensions[row].height = 32
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)


def _autosize(sheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def payroll_workbook(results: list[PayrollResult], period: str) -> bytes:
    """Oylik hisob-kitobi: bitta jamlovchi varaq + har bir xodim uchun tafsilot."""
    wb = Workbook()
    summary = wb.active
    summary.title = "Oylik jamlanma"

    summary["A1"] = f"Oylik hisob-kitobi — {period}"
    summary["A1"].font = Font(bold=True, size=14)
    summary.merge_cells("A1:N1")

    titles = [
        "№",
        "F.I.Sh.",
        "Oylik (so'm)",
        "Ish kuni (norma)",
        "Ishlagan kun",
        "Kelmagan kun",
        "Ta'til/ruxsat",
        "Kechikish (marta)",
        "Kechikish (jami daq)",
        "Jarima ostidagi daq",
        "Kechikish jarimasi",
        "Erta ketish (daq)",
        "Tushlik oshig'i (daq)",
        "Ishlanmagan vaqt ushlanmasi",
        "Ustama kamayishi",
        "To'lanadi (so'm)",
    ]
    _write_header(summary, 3, titles)

    row = 4
    for index, res in enumerate(results, start=1):
        values = [
            index,
            res.employee_name,
            float(res.monthly_salary),
            res.scheduled_days,
            res.worked_days,
            res.absent_days,
            res.leave_days,
            res.late_count,
            res.total_late_minutes,
            res.total_fined_late_minutes,
            float(res.late_fine),
            res.total_early_leave_minutes,
            res.total_lunch_overrun_minutes,
            float(res.unpaid_time_deduction),
            float(res.bonus_reduction),
            float(res.net_payable),
        ]
        for col, value in enumerate(values, start=1):
            cell = summary.cell(row=row, column=col, value=value)
            cell.border = _BORDER
            if col in (3, 11, 14, 15, 16):
                cell.number_format = _MONEY_FMT
        row += 1

    # Jami satri
    total_row = row
    summary.cell(row=total_row, column=2, value="JAMI").font = _TOTAL_FONT
    for col in (3, 11, 14, 15, 16):
        letter = get_column_letter(col)
        cell = summary.cell(
            row=total_row, column=col, value=f"=SUM({letter}4:{letter}{row - 1})"
        )
        cell.font = _TOTAL_FONT
        cell.number_format = _MONEY_FMT
        cell.border = _BORDER

    _autosize(summary, [4, 28, 15, 12, 12, 12, 12, 12, 14, 14, 17, 13, 15, 20, 16, 18])

    # Har bir xodim uchun kunlik tafsilot
    for res in results:
        if not res.days:
            continue
        name = res.employee_name[:28] or f"Xodim {res.employee_id}"
        sheet = wb.create_sheet(title=name.replace("/", "-").replace("\\", "-"))
        sheet["A1"] = f"{res.employee_name} — {period}"
        sheet["A1"].font = Font(bold=True, size=12)
        sheet.merge_cells("A1:G1")
        _write_header(
            sheet,
            3,
            [
                "Sana",
                "Hafta kuni",
                "Holat",
                "Kechikish (daq)",
                "Erta ketish (daq)",
                "Tushlik oshig'i (daq)",
                "To'lanmaydigan (daq)",
                "Izoh",
            ],
        )
        detail_row = 4
        for item in res.days:
            values = [
                item.day.strftime("%d.%m.%Y"),
                UZ_WEEKDAYS[item.day.isoweekday()],
                _STATUS_LABELS.get(item.status, item.status.value),
                item.late_minutes,
                item.early_leave_minutes,
                item.lunch_overrun_minutes,
                item.unpaid_minutes,
                item.note,
            ]
            fill = _STATUS_FILLS.get(item.status)
            for col, value in enumerate(values, start=1):
                cell = sheet.cell(row=detail_row, column=col, value=value)
                cell.border = _BORDER
                if fill is not None:
                    cell.fill = fill
            detail_row += 1
        _autosize(sheet, [12, 14, 14, 15, 16, 18, 18, 26])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell_for_day(
    db: Session, employee: Employee, day: date, record: Attendance | None
) -> tuple[str, DayStatus | None]:
    """Tabelning bitta katagi: matn va bo'yash uchun holat.

    Yozuv bo'lmasa ham to'g'ri belgi qo'yiladi: dam olish kunlari va ta'til
    uchun `attendance` qatori umuman yaratilmaydi (skanerlash bo'lmagan),
    shuning uchun holatni shu yerda hisoblaymiz — aks holda bayram va ta'til
    tabelda bo'sh katak bo'lib qolardi.
    """
    if record is not None:
        if record.status == DayStatus.HOLIDAY:
            return "D", DayStatus.HOLIDAY
        if record.status == DayStatus.LEAVE:
            return "T", DayStatus.LEAVE
        if record.check_in_at is None:
            return "—", record.status
        return fmt_time(record.check_in_at), record.status

    if not employee.schedule.is_work_day(day) or att_mod.is_holiday(db, day) is not None:
        return "D", DayStatus.HOLIDAY
    if att_mod.leave_for(db, employee.id, day) is not None:
        return "T", DayStatus.LEAVE
    return "—", DayStatus.ABSENT


def attendance_workbook(db: Session, period: str) -> bytes:
    """Oylik davomat tabeli: qatorlar — xodimlar, ustunlar — kunlar."""
    start, end = month_bounds(period)
    days = list(iter_days(start, end))

    employees = db.scalars(
        select(Employee).where(Employee.is_active.is_(True)).order_by(Employee.full_name)
    ).all()

    records: dict[tuple[int, date], Attendance] = {
        (row.employee_id, row.work_date): row
        for row in db.scalars(
            select(Attendance).where(
                Attendance.work_date >= start, Attendance.work_date <= end
            )
        ).all()
    }

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Tabel"
    sheet["A1"] = f"Davomat tabeli — {period}"
    sheet["A1"].font = Font(bold=True, size=14)

    _write_header(sheet, 3, ["F.I.Sh."] + [d.strftime("%d") for d in days])
    for col, day in enumerate(days, start=2):
        sheet.cell(row=2, column=col, value=UZ_WEEKDAYS[day.isoweekday()][:2]).alignment = (
            Alignment(horizontal="center")
        )

    row = 4
    for emp in employees:
        sheet.cell(row=row, column=1, value=emp.full_name).border = _BORDER
        for col, day in enumerate(days, start=2):
            text, status = _cell_for_day(db, emp, day, records.get((emp.id, day)))
            cell = sheet.cell(row=row, column=col, value=text)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
            fill = _STATUS_FILLS.get(status) if status is not None else None
            if fill is not None:
                cell.fill = fill
        row += 1

    _autosize(sheet, [28] + [7] * len(days))

    legend = row + 1
    legend_text = (
        "Izoh: soat = kelgan vaqti,  D = dam olish,  T = ta'til/ruxsat,  — = kelmagan"
    )
    sheet.cell(row=legend, column=1, value=legend_text)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
