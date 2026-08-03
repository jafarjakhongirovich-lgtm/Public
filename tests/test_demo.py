"""Demo ma'lumot va bayram taqvimi testlari."""

from __future__ import annotations

from datetime import date

import pytest
from sqlalchemy import select

from app import demo as demo_mod
from app import holidays_uz
from app import payroll as payroll_mod
from app.models import Attendance, AttendanceEvent, DayStatus, Employee, Holiday, Kiosk, Leave

PERIOD = "2026-07"


# --------------------------------------------------------------------------- #
# Demo ma'lumot
# --------------------------------------------------------------------------- #


def test_seed_demo_creates_employees_and_events(db):
    result = demo_mod.seed_demo(db, PERIOD)

    assert result["employees"] == len(demo_mod.PROFILES)
    assert result["events"] > 300  # 6 xodim x ~22 kun x ~4 yozuv
    assert db.scalars(select(AttendanceEvent)).first() is not None
    assert db.scalars(select(Kiosk)).first() is not None


def test_seed_demo_is_deterministic(db):
    """Bir xil seed — bir xil natija. Aks holda testlar ishonchsiz bo'lardi."""
    first = demo_mod.seed_demo(db, PERIOD, seed=42)
    totals_first = sorted(
        (row.employee_id, row.work_date, row.check_in_at)
        for row in db.scalars(select(Attendance)).all()
    )

    # Bazani tozalab, xuddi shu seed bilan qaytadan
    for model in (Attendance, AttendanceEvent, Leave, Holiday, Employee):
        for row in db.scalars(select(model)).all():
            db.delete(row)
    db.flush()

    second = demo_mod.seed_demo(db, PERIOD, seed=42, force=True)
    totals_second = sorted(
        (row.employee_id, row.work_date, row.check_in_at)
        for row in db.scalars(select(Attendance)).all()
    )

    assert first["events"] == second["events"]
    # employee_id lar yangi bo'ladi, shuning uchun faqat vaqtlarni solishtiramiz
    assert [t[2] for t in totals_first] == [t[2] for t in totals_second]


def test_seed_demo_refuses_to_overwrite(db, employee):
    """Haqiqiy xodimlar ustiga demo yozib qo'yish — xavfli, rad etiladi."""
    with pytest.raises(demo_mod.DemoDataExists):
        demo_mod.seed_demo(db, PERIOD)


def test_seed_demo_force_allows_existing(db, employee):
    result = demo_mod.seed_demo(db, PERIOD, force=True)
    assert result["employees"] == len(demo_mod.PROFILES)


def test_demo_metrics_come_from_real_code(db):
    """Demo yozuvlari haqiqiy hisoblash yo'lidan o'tishi kerak — statuslar,
    kechikish va ishlangan vaqt qo'lda "chizilmaydi"."""
    demo_mod.seed_demo(db, PERIOD)

    records = db.scalars(select(Attendance)).all()
    assert records, "kunlik xulosalar yaratilmagan"

    # Har bir to'liq kunda ishlangan vaqt hisoblangan bo'lishi kerak
    complete = [r for r in records if r.check_in_at and r.check_out_at]
    assert complete
    assert all(r.worked_minutes > 0 for r in complete)

    # Kamida bitta kechikish va bitta tushlik oshig'i bo'lsin (profillar shunga sozlangan)
    assert any(r.billable_late_minutes > 0 for r in records)
    assert any(r.lunch_overrun_minutes > 0 for r in records)


def test_demo_produces_varied_statuses(db):
    """Demo turli holatlarni ko'rsatishi kerak — aks holda panelni sinash befoyda."""
    demo_mod.seed_demo(db, PERIOD)
    statuses = {row.status for row in db.scalars(select(Attendance)).all()}

    assert DayStatus.LATE in statuses
    assert DayStatus.PRESENT in statuses
    # Ketishini qayd etmagan kunlar ham bo'lishi kerak (forgets_checkout_chance)
    assert DayStatus.INCOMPLETE in statuses

    # Bayram va ta'til kunlarida skanerlash bo'lmaydi, shuning uchun `attendance`
    # qatori ham yaratilmaydi — ular tabelda hisoblab chiqariladi (test pastda).
    assert db.scalars(select(Holiday)).first() is not None
    assert db.scalars(select(Leave)).first() is not None


def test_timesheet_marks_holidays_and_leave_without_records(db):
    """Bayram va ta'til kunlari tabelda bo'sh QOLMASLIGI kerak.

    Bu kunlarda hech kim skanerlamaydi, ya'ni `attendance` qatori yo'q. Tabel
    holatni jadval/bayram/ta'til jadvallaridan hisoblab chiqarishi kerak.
    """
    import io

    from openpyxl import load_workbook

    from app import excel as excel_mod

    demo_mod.seed_demo(db, PERIOD)
    sheet = load_workbook(io.BytesIO(excel_mod.attendance_workbook(db, PERIOD))).active

    # 4-qatordan boshlab xodimlar, 2-ustundan boshlab kunlar
    values = {
        sheet.cell(row=r, column=c).value
        for r in range(4, 4 + len(demo_mod.PROFILES))
        for c in range(2, 2 + 31)
    }
    assert "D" in values, "dam olish/bayram kunlari belgilanmagan"
    assert "T" in values, "ta'til kunlari belgilanmagan"
    # Bo'sh katak qolmasligi kerak (oyning oxirida ustun bo'lmagan joylardan tashqari)
    assert None not in {
        sheet.cell(row=4, column=c).value for c in range(2, 2 + 31)
    }, "tabelda bo'sh kataklar bor"


def test_demo_payroll_is_sane(db):
    """Har bir xodimga to'lov 0 dan katta va oylikdan oshmasligi kerak."""
    demo_mod.seed_demo(db, PERIOD)
    results = payroll_mod.compute_period(db, PERIOD)

    assert len(results) == len(demo_mod.PROFILES)
    for res in results:
        assert res.scheduled_days > 0
        assert 0 < res.net_payable <= res.monthly_salary, res.employee_name
        assert res.accrued <= res.monthly_salary

    # Muntazam kechiqadigan xodim to'liq oylik olmasligi kerak
    latecomer = next(r for r in results if r.employee_name.startswith("Rasulov"))
    assert latecomer.net_payable < latecomer.monthly_salary
    assert latecomer.late_count > 0

    # Vaqtida keladigan xodim to'liq oylik olishi kerak
    punctual = next(r for r in results if r.employee_name.startswith("Aliyev"))
    assert punctual.net_payable == punctual.monthly_salary


def test_demo_employees_have_no_telegram_link(db):
    """Demo xodimlar Telegram'ga bog'lanmaydi — bot bo'lmasa ham ishlashi uchun,
    va haqiqiy odamning user_id si tasodifan band bo'lib qolmasligi uchun."""
    demo_mod.seed_demo(db, PERIOD)
    employees = db.scalars(select(Employee)).all()
    assert all(e.telegram_user_id is None for e in employees)


def test_demo_excel_export_works(db):
    from app import excel as excel_mod

    demo_mod.seed_demo(db, PERIOD)
    results = payroll_mod.compute_period(db, PERIOD)

    assert excel_mod.payroll_workbook(results, PERIOD)[:2] == b"PK"
    assert excel_mod.attendance_workbook(db, PERIOD)[:2] == b"PK"


# --------------------------------------------------------------------------- #
# Bayram taqvimi
# --------------------------------------------------------------------------- #


def test_fixed_holidays_for_year():
    days = holidays_uz.fixed_for_year(2026)
    assert len(days) == len(holidays_uz.FIXED_HOLIDAYS)
    assert (date(2026, 1, 1), "Yangi yil") in days
    assert (date(2026, 9, 1), "Mustaqillik kuni") in days
    assert all(day.year == 2026 for day, _ in days)


def test_fixed_holidays_shift_with_year():
    for year in (2025, 2026, 2027, 2030):
        assert all(day.year == year for day, _ in holidays_uz.fixed_for_year(year))


def test_lunar_holidays_are_not_treated_as_certain():
    """Hayit sanalari taxminiy — nomida shu aytilgan bo'lishi kerak, chunki
    ular avtomatik qo'shilsa oylik xato hisoblanishi mumkin."""
    hints = holidays_uz.lunar_hints_for_year(2026)
    assert hints
    assert all("taxminiy" in name for _, name in hints)


def test_unknown_year_returns_no_lunar_hints():
    assert holidays_uz.lunar_hints_for_year(2099) == []


def test_weekend_collisions_detected():
    """2026: 8 mart — yakshanba, 21 mart va 9 may — shanba."""
    collisions = dict(holidays_uz.weekend_collisions(2026))
    assert date(2026, 3, 8) in collisions
    assert date(2026, 3, 21) in collisions
    assert date(2026, 5, 9) in collisions
    # 1 yanvar 2026 — payshanba, ro'yxatga tushmasligi kerak
    assert date(2026, 1, 1) not in collisions


def test_holidays_reduce_payroll_norm(db, employee):
    """Bayram qo'shilgandan keyin oylik normasi kamayishi kerak."""
    before = payroll_mod.compute_employee(db, employee, PERIOD).scheduled_days

    for day, name in holidays_uz.fixed_for_year(2026):
        if day.month == 7:  # iyulda qat'iy bayram yo'q, shuning uchun qo'shamiz
            db.add(Holiday(day=day, name=name))
    db.add(Holiday(day=date(2026, 7, 20), name="Sinov bayrami"))
    db.flush()

    after = payroll_mod.compute_employee(db, employee, PERIOD).scheduled_days
    assert after == before - 1
